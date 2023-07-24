from openpyxl import load_workbook


# Class for the excel file controls
class ExcelControl:
    def __init__(self,excel_path,sheet_name):
        self.excel_path = excel_path # Excel file path to control
        self.sheet_name = sheet_name # Excel Sheet name to control
        self.wb = load_workbook(self.excel_path, data_only=True) # Get excel workbook in data only mode
        self.sh = self.wb[self.sheet_name] # Get excel sheet
        


    def read_from_excel(self,cell : str):
        """
        It reads value from a specific cell in the excel file sheet.\n
        Arg: cell -> the cell value has to be 'LetterInteger' format.\n
        Example : "A1","B14"
        """
        return self.sh[cell].value # Return value of the cell provided

    def add_to_excel(self, cell : str, value : str):
        """
        It adds value to a specific cell in the excel file sheet.\n
        Arg: cell -> the cell value has to be 'LetterInteger' format.\n
             value -> the value to be added to the cell 
        Example : cell:"A1",value:"3"
        """
        self.sh[cell].value = value # Assign value provided to the cell provided
        self.wb.save(self.excel_path) # Finally, save the excel file

    def get_row_count_from_excel(self):
        """
        It returns row count of the excel file sheet
        """
        
        return self.sh.max_row # Return row count of the  excel sheet
    
    def delete_from_excel(self, row_count: int):
        """
        It deletes row given from the excel
        """
        self.sh.delete_rows(row_count) # Delete the row provided from the excel
        self.wb.save(self.excel_path)  # Finally, save the excel file